from openpyxl.worksheet.worksheet import Worksheet
from src.typeDefs.fileInfo import IFileInfo
import datetime as dt
from openpyxl import load_workbook
import os
from typing import List
from src.config.appConfig import getMeasInfo
from datetime import timedelta
from src.typeDefs.metricsDataRecord import IMetricsDataRecord

def getFileData(fileInfo: IFileInfo, targetDt: dt.datetime) -> List[IMetricsDataRecord]:
    # TODO fetch with openpyxl
    # convert datetime to string
    targetDtFormatStr = dt.datetime.strftime(targetDt, '%Y-%m-%d')
    # convert datetime to datetime format
    targetDt = dt.datetime.strptime(targetDtFormatStr, '%Y-%m-%d')
    fileDateStr = dt.datetime.strftime(targetDt,fileInfo['date_format'])
    targetFilename = fileInfo['filename'].replace('{{dt}}', fileDateStr)
    targetFilePath = os.path.join(fileInfo['folder_location'], targetFilename)
    wb = load_workbook(targetFilePath, data_only=True)
    ws = wb.active
    
    measInfo = getMeasInfo()
    records:List[IMetricsDataRecord] = []
    for eachMeasInfo in measInfo:
        if fileInfo['file_tag'] == eachMeasInfo['file_tag']:
            time_offset = eachMeasInfo['time_offset_hrs_mins'].split('_')
            address = eachMeasInfo['address']
            entity_tag = eachMeasInfo['entity_tag']
            metric_name = eachMeasInfo['metric_name']
            time_stamp = targetDt + timedelta(hours= int(time_offset[0]), minutes=int(time_offset[1]))
            data_val = ws[address].value

            record:IMetricsDataRecord = {
                "data_time": time_stamp,
                "entity_tag": entity_tag,
                "metric_name": metric_name,
                "data_val": data_val
            }
            records.append(record)

    return records
